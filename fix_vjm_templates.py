"""
fix_vjm_templates.py - Ajusta las plantillas VJM por tipo:
- Expande ImagenesComponentes con filas correctas (agencies + favoriteCities + carRental)
- Corrige orden de componente (fix bug template agencia F3=F4=1 -> 1,2)

Por tipo:
  CIUDAD: 2 agencies + 16 favoriteCities + 5 carRental = 23 filas
  AGENCIA: 2 agencies + 16 favoriteCities + 5 carRental = 23 filas
  LOCALIDAD: 2 agencies + 15 favoriteCities = 17 filas (sin carRental)
  TIPO_AUTO: 2 agencies + 15 favoriteCities + 5 carRental = 22 filas
"""
import os
from copy import copy
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(ROOT, "RESULTADOS VJS", "CARGUES DE CONTENIDO VJS")

CONFIG = {
    "ciudad":    {"agencies": 2, "favoriteCities": 16, "carRental": 5},
    "agencia":   {"agencies": 2, "favoriteCities": 15, "carRental": 5},
    "localidad": {"agencies": 2, "favoriteCities": 15, "carRental": 0},
    "tipo_auto": {"agencies": 2, "favoriteCities": 15, "carRental": 5},
}


def copy_row_style(src_cell, dst_cell):
    """Copia el estilo de src_cell a dst_cell."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def fix_imagenes_componentes(ws, config):
    """Expande ImagenesComponentes segun config."""
    # Limpiar filas 3-15 (las existentes)
    for row in range(3, 20):
        for col in range(1, 13):
            ws.cell(row=row, column=col, value="")

    # Escribir nuevas filas segun config
    row = 3
    for seccion in ("agencies", "favoriteCities", "carRental"):
        count = config.get(seccion, 0)
        for orden in range(1, count + 1):
            ws.cell(row=row, column=1, value=seccion)
            ws.cell(row=row, column=2, value=orden)
            # Aplicar estilo de fila 3 (la base original)
            if row > 3:
                for col in range(1, 13):
                    src = ws.cell(row=3, column=col)
                    dst = ws.cell(row=row, column=col)
                    copy_row_style(src, dst)
            row += 1


def main():
    for tipo, config in CONFIG.items():
        path = os.path.join(TEMPLATES_DIR, f"loadContentVjs_{tipo}.xlsx")
        if not os.path.exists(path):
            print(f"NO existe: {path}")
            continue
        print(f"Fixing {tipo}: {sum(config.values())} filas imagenes componentes")
        wb = openpyxl.load_workbook(path)
        fix_imagenes_componentes(wb["ImagenesComponentes"], config)
        wb.save(path)
        print(f"  OK: {path}")


if __name__ == "__main__":
    main()
