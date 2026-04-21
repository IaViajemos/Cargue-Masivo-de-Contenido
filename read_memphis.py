import openpyxl
import os

fpath = r"c:\Users\Miguel Martinez SSD\OneDrive - BROWSER TRAVEL SOLUTIONS S.A.S VIAJEMOS\Documentos\PROYECTOS\CARGUE MASIVO VJM Y MCR\LANDINGS EJEMPLO MCR\CIUDADES\09-01-2025 Cargue de Contenido Memphis.xlsx"

wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
ws = wb["Secciones"]

def get_val(cell):
    if cell is None:
        return ""
    v = cell
    if v is None:
        return ""
    return str(v).strip()

def word_count(text):
    if not text:
        return 0
    return len(text.split())

print("=" * 120)
print("SECTION 1: FLEETCARRUSEL (rows 28-39)")
print("=" * 120)

for row in ws.iter_rows(min_row=28, max_row=39, min_col=1, max_col=8, values_only=False):
    rn = row[0].row
    col_a = get_val(row[0].value)
    col_b = get_val(row[1].value)
    col_c = get_val(row[2].value)
    col_d = get_val(row[3].value)
    col_e = get_val(row[4].value)
    col_f = get_val(row[5].value)
    col_g = get_val(row[6].value)
    col_h = get_val(row[7].value)

    wc_c = word_count(col_c)
    wc_d = word_count(col_d)
    wc_e = word_count(col_e)

    print(f"\n--- ROW {rn} ---")
    if col_a:
        print(f"  Col A (Seccion):      {col_a}")
    print(f"  Col B (Comentarios):  {col_b}")
    print(f"  Col F (Tipo formato): {col_f}")
    if col_g:
        print(f"  Col G (ID Section):   {col_g}")
    if col_h:
        print(f"  Col H (ID Component): {col_h}")
    print(f"  Col C (Espanol) [{wc_c} words]:")
    print(f"    {col_c}")
    print(f"  Col D (English) [{wc_d} words]:")
    print(f"    {col_d}")
    print(f"  Col E (Portugues) [{wc_e} words]:")
    print(f"    {col_e}")

print("\n\n")
print("=" * 120)
print("SECTION 2: LOCATIONSCARRUSEL (rows 40-73)")
print("=" * 120)

for row in ws.iter_rows(min_row=40, max_row=73, min_col=1, max_col=8, values_only=False):
    rn = row[0].row
    col_a = get_val(row[0].value)
    col_b = get_val(row[1].value)
    col_c = get_val(row[2].value)
    col_d = get_val(row[3].value)
    col_e = get_val(row[4].value)
    col_f = get_val(row[5].value)
    col_g = get_val(row[6].value)
    col_h = get_val(row[7].value)

    wc_c = word_count(col_c)
    wc_d = word_count(col_d)
    wc_e = word_count(col_e)

    print(f"\n--- ROW {rn} ---")
    if col_a:
        print(f"  Col A (Seccion):      {col_a}")
    print(f"  Col B (Comentarios):  {col_b}")
    print(f"  Col F (Tipo formato): {col_f}")
    if col_g:
        print(f"  Col G (ID Section):   {col_g}")
    if col_h:
        print(f"  Col H (ID Component): {col_h}")
    print(f"  Col C (Espanol) [{wc_c} words]:")
    print(f"    {col_c}")
    print(f"  Col D (English) [{wc_d} words]:")
    print(f"    {col_d}")
    print(f"  Col E (Portugues) [{wc_e} words]:")
    print(f"    {col_e}")

wb.close()
print("\n\nDONE")
