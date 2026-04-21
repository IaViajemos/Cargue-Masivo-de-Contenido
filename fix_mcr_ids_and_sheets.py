"""
fix_mcr_ids_and_sheets.py — Corrige IDs y hojas auxiliares en templates MCR.

Para cada template MCR:
1. Recalcula ID Component (col H) secuencialmente en Secciones
2. Actualiza ImagenesComponentes con las 17 ciudades estándar
3. Actualiza BotonComponentes para cubrir todos los ID Component
4. Actualiza GaleriasComponentes con el ID correcto del blog
"""
import sys
import os
from copy import copy
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = os.path.dirname(os.path.abspath(__file__))

# 17 ciudades estándar para ImagenesComponentes cluster
CIUDADES_17 = [
    ("miami", "Miami"),
    ("orlando", "Orlando"),
    ("cbx", "CBX"),
    ("las-vegas", "Las Vegas"),
    ("nueva-york", "Nueva York"),
    ("los-angeles", "Los Ángeles"),
    ("houston", "Houston"),
    ("chicago", "Chicago"),
    ("fort-lauderdale", "Fort Lauderdale"),
    ("san-diego", "San Diego"),
    ("dallas", "Dallas"),
    ("phoenix", "Phoenix"),
    ("tampa", "Tampa"),
    ("san-francisco", "San Francisco"),
    ("atlanta", "Atlanta"),
    ("denver", "Denver"),
    ("austin", "Austin"),
]

# 5 tipos de flota
FLEET_TYPES = [
    (2, "economicos-mcr", "Carros Economicos"),
    (3, "camionetas-mcr", "Camionetas"),
    (4, "vans-mcr", "Vans"),
    (5, "convertibles-mcr", "Convertibles"),
    (6, "lujo-mcr", "Carros de Lujo"),
    (7, "electricos-mcr", "Carros Electricos"),  # solo tipo_auto
]

TEMPLATES = {
    "ciudad": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr.xlsx"),
    "agencia": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr agencias.xlsx"),
    "localidad": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr localidades.xlsx"),
    "tipo_auto": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr tipos autos.xlsx"),
}


def recalc_ids(ws):
    """Recalcula ID Component (col H) secuencialmente.

    Reglas:
    - Disclaimer de questions SIEMPRE = 33
    - Disclaimer de rentcompanies = 1
    - Todo lo demás: secuencial desde 2
    - Se asigna ID solo a: Descripción H3, Disclaimer, text_end
    """
    # Primero limpiar todos los H existentes (excepto header)
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=8, value=None)

    # Identificar bloques y asignar IDs
    current_block = None
    next_id = 2  # empezamos en 2 (1 reservado para disclaimer rentcompanies)
    disclaimer_q_row = None
    disclaimer_rc_row = None

    # Primer pase: encontrar disclaimers
    for row in range(3, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        b = (ws.cell(row, 2).value or '').strip()
        if a:
            current_block = a
        if b in ('Disclaimer', 'Disclaimer F'):
            if current_block in ('questions', 'rentalCarFaqs'):
                disclaimer_q_row = row
            elif current_block == 'rentcompanies':
                disclaimer_rc_row = row
            elif current_block in ('text_end_landingpage',):
                pass  # text_end disclaimer se numera secuencial

    # Segundo pase: asignar IDs secuenciales
    current_block = None
    for row in range(3, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        b = (ws.cell(row, 2).value or '').strip()

        if a:
            current_block = a

        # Disclaimer rentcompanies → ID 1 (fijo)
        if row == disclaimer_rc_row:
            ws.cell(row=row, column=8, value=1)
            continue

        # Disclaimer questions → ID 33 (fijo)
        if row == disclaimer_q_row:
            ws.cell(row=row, column=8, value=33)
            continue

        # Descripción H3 → ID secuencial
        if b.startswith('Descripción H3'):
            ws.cell(row=row, column=8, value=next_id)
            next_id += 1

    # text_end_landingpage → último ID secuencial
    for row in range(3, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        if 'text_end' in a:
            ws.cell(row=row, column=8, value=next_id)
            next_id += 1
            break

    return next_id  # total de componentes + 1


def get_all_comp_ids(ws):
    """Extrae todos los ID Component asignados (col H)."""
    ids = []
    for row in range(3, ws.max_row + 1):
        h = ws.cell(row, 8).value
        if h and isinstance(h, int) and h not in (1, 33):
            ids.append(h)
    return sorted(ids)


def update_imagenes_componentes(wb, tipo, comp_ids):
    """Actualiza ImagenesComponentes con fleet + 17 ciudades cluster."""
    ws = wb['ImagenesComponentes']

    # Limpiar desde fila 3
    for r in range(3, max(ws.max_row + 1, 30)):
        for c in range(1, 16):
            ws.cell(row=r, column=c, value=None)

    row = 3

    # Determinar qué componentes son fleet y cuáles locations
    # Basado en el orden de bloques en Secciones
    ws_sec = wb['Secciones']
    fleet_comp_ids = []
    loc_comp_ids = []
    rentacar_comp_ids = []
    current_block = None

    for r in range(3, ws_sec.max_row + 1):
        a = ws_sec.cell(r, 1).value or ''
        b = (ws_sec.cell(r, 2).value or '').strip()
        h = ws_sec.cell(r, 8).value

        if a:
            current_block = a
        if h and isinstance(h, int) and h not in (1, 33):
            if current_block == 'fleetcarrusel':
                fleet_comp_ids.append(h)
            elif current_block == 'locationscarrusel':
                loc_comp_ids.append(h)
            elif current_block == 'rentacar':
                rentacar_comp_ids.append(h)

    # Fleet images (5 o 6 tipos)
    fleet_count = min(len(fleet_comp_ids), len(FLEET_TYPES))
    for i in range(fleet_count):
        comp_id = fleet_comp_ids[i]
        _, img_slug, label = FLEET_TYPES[i]
        jpg = f"images/ContentLp/{img_slug}.jpg"
        webp = f"images/ContentLp/{img_slug}.webp"

        ws.cell(row=row, column=1, value=comp_id)
        ws.cell(row=row, column=2, value=jpg)
        ws.cell(row=row, column=3, value=webp)
        ws.cell(row=row, column=4, value=f"Renta de autos {label.lower()}")
        ws.cell(row=row, column=5, value=f"Alquiler de carros {label.lower()}")
        ws.cell(row=row, column=6, value=None)  # HREF se llena en builder
        ws.cell(row=row, column=7, value=f"Rent {label.lower()} cars")
        ws.cell(row=row, column=8, value=f"Affordable {label.lower()} car rentals")
        ws.cell(row=row, column=9, value=None)
        ws.cell(row=row, column=10, value=f"Aluguel de carros {label.lower()}")
        ws.cell(row=row, column=11, value=f"Locação de veículos {label.lower()}")
        ws.cell(row=row, column=12, value=None)
        row += 1

    # Location cluster images (17 ciudades)
    for i, comp_id in enumerate(loc_comp_ids):
        if i < len(CIUDADES_17):
            slug, name = CIUDADES_17[i]
        else:
            slug, name = f"ciudad-{i+1}", f"Ciudad {i+1}"
        jpg = f"images/ContentLp/{slug}-mcr-cluster.jpg"
        webp = f"images/ContentLp/{slug}-mcr-cluster.webp"

        ws.cell(row=row, column=1, value=comp_id)
        ws.cell(row=row, column=2, value=jpg)
        ws.cell(row=row, column=3, value=webp)
        ws.cell(row=row, column=4, value=f"Renta de autos en {name}")
        ws.cell(row=row, column=5, value=f"Alquiler de autos en {name}")
        ws.cell(row=row, column=6, value=None)
        ws.cell(row=row, column=7, value=f"Car rentals in {name}")
        ws.cell(row=row, column=8, value=f"Car rental in {name}")
        ws.cell(row=row, column=9, value=None)
        ws.cell(row=row, column=10, value=f"Aluguel de carros em {name}")
        ws.cell(row=row, column=11, value=f"Locação de veículos em {name}")
        ws.cell(row=row, column=12, value=None)
        row += 1

    return row - 3  # total filas escritas


def update_boton_componentes(wb):
    """Actualiza BotonComponentes con todos los ID Component."""
    ws_sec = wb['Secciones']
    ws_btn = wb['BotonComponentes']

    # Recoger todos los IDs de componentes (excepto 1 y 33)
    all_ids = []
    for r in range(3, ws_sec.max_row + 1):
        h = ws_sec.cell(r, 8).value
        if h and isinstance(h, int) and h not in (1, 33):
            all_ids.append(h)
    all_ids = sorted(set(all_ids))

    # Limpiar desde fila 3
    for r in range(3, max(ws_btn.max_row + 1, 30)):
        for c in range(1, 5):
            ws_btn.cell(row=r, column=c, value=None)

    # Escribir un botón por cada ID
    for i, comp_id in enumerate(all_ids):
        row = 3 + i
        ws_btn.cell(row=row, column=1, value=comp_id)
        ws_btn.cell(row=row, column=2, value="Ver Ofertas")
        ws_btn.cell(row=row, column=3, value="See Offers")
        ws_btn.cell(row=row, column=4, value="Ver Ofertas")

    return len(all_ids)


def update_galerias_componentes(wb):
    """Actualiza GaleriasComponentes: blog apunta al text_end_landingpage ID."""
    ws_sec = wb['Secciones']
    ws_gal = wb['GaleriasComponentes']

    # Buscar el ID de text_end_landingpage
    blog_id = None
    for r in range(3, ws_sec.max_row + 1):
        a = ws_sec.cell(r, 1).value or ''
        h = ws_sec.cell(r, 8).value
        if 'text_end' in a and h:
            blog_id = h
            break

    if blog_id:
        ws_gal.cell(row=3, column=2, value=blog_id)

    return blog_id


def main():
    for tipo, path in TEMPLATES.items():
        try:
            wb = openpyxl.load_workbook(path)
        except PermissionError:
            print(f"  {tipo}: BLOQUEADO")
            continue

        print(f"\n{'='*50}")
        print(f"  {tipo.upper()}")
        print(f"{'='*50}")

        # 1. Recalcular IDs en Secciones
        ws = wb['Secciones']
        next_id = recalc_ids(ws)
        comp_ids = get_all_comp_ids(ws)
        print(f"  Secciones: {len(comp_ids)} IDs asignados (2-{max(comp_ids) if comp_ids else 0})")

        # 2. Actualizar ImagenesComponentes
        img_count = update_imagenes_componentes(wb, tipo, comp_ids)
        print(f"  ImagenesComponentes: {img_count} filas")

        # 3. Actualizar BotonComponentes
        btn_count = update_boton_componentes(wb)
        print(f"  BotonComponentes: {btn_count} botones")

        # 4. Actualizar GaleriasComponentes
        blog_id = update_galerias_componentes(wb)
        print(f"  GaleriasComponentes: blog_id={blog_id}")

        wb.save(path)
        print(f"  Guardado OK")


if __name__ == "__main__":
    main()
