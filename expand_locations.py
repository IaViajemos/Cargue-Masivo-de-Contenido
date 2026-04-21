"""
expand_locations.py — Expande los bloques favoriteCities/locationscarrusel
a 17 H3 slots en TODAS las plantillas MCR y VJM.

Para cada plantilla:
1. Encuentra el bloque (locationscarrusel o favoriteCities)
2. Calcula cuántos H3 faltan para llegar a 17
3. Inserta filas (pares H3+desc) al final del bloque
4. Copia estilos de las filas existentes
5. Rellena col A/B/F de las nuevas filas
6. Guarda

Esto desplaza las filas posteriores automáticamente (openpyxl.insert_rows).
"""
import sys
import openpyxl
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

TEMPLATES = {
    "MCR CIUDAD": {
        "path": "RESULTADOS MCR/CARGUES DE CONTENIDO MCR/loadContentMcr.xlsx",
        "bloque": "locationscarrusel",
    },
    "MCR LOCALIDAD": {
        "path": "RESULTADOS MCR/CARGUES DE CONTENIDO MCR/loadContentMcr localidades.xlsx",
        "bloque": "locationscarrusel",
    },
    "VJM CIUDAD": {
        "path": "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_ciudad.xlsx",
        "bloque": "favoriteCities",
    },
    "VJM AGENCIA": {
        "path": "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_agencia.xlsx",
        "bloque": "favoriteCities",
    },
    "VJM LOCALIDAD": {
        "path": "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_localidad.xlsx",
        "bloque": "favoriteCities",
    },
    "VJM TIPO_AUTO": {
        "path": "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_tipo_auto.xlsx",
        "bloque": "favoriteCities",
    },
}

TARGET_H3 = 17


def find_block_range(ws, bloque_name):
    """Encuentra el rango de filas del bloque y cuenta H3."""
    in_blk = False
    start = end = 0
    h3_count = 0
    last_h3_row = 0
    for row in range(3, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        b = (ws.cell(row, 2).value or '').strip()
        if bloque_name in a:
            in_blk = True
            start = row
        elif a and bloque_name not in a and in_blk:
            end = row - 1
            break
        if in_blk and b == 'H3':
            h3_count += 1
            last_h3_row = row
    if not end:
        end = ws.max_row
    return start, end, h3_count, last_h3_row


def copy_row_style(ws, src_row, dst_row, max_col=11):
    """Copia estilos de src_row a dst_row."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def expand_block(ws, bloque_name, target=17):
    """Inserta filas para que el bloque tenga exactamente target H3."""
    start, end, h3_count, last_h3_row = find_block_range(ws, bloque_name)
    if h3_count >= target:
        return 0  # ya tiene suficientes

    need = target - h3_count
    # La ultima desc esta en last_h3_row + 1
    insert_after = last_h3_row + 1  # insertar DESPUES de la ultima desc
    insert_point = insert_after + 1  # fila donde insertar

    # Insertar filas (2 por cada H3 faltante: H3 + desc)
    rows_to_insert = need * 2
    ws.insert_rows(insert_point, rows_to_insert)

    # Llenar las nuevas filas con estructura H3/desc
    # Buscar una fila H3 existente para copiar estilo
    style_h3_row = last_h3_row
    style_desc_row = last_h3_row + 1

    for i in range(need):
        h3_row = insert_point + i * 2
        desc_row = insert_point + i * 2 + 1

        # Copiar estilos
        copy_row_style(ws, style_h3_row, h3_row)
        copy_row_style(ws, style_desc_row, desc_row)

        # Llenar col B (etiqueta) y col F (tipo formato)
        ws.cell(row=h3_row, column=2, value="H3")
        ws.cell(row=h3_row, column=6, value="component_title")
        ws.cell(row=desc_row, column=2, value="Descripción H3")
        ws.cell(row=desc_row, column=6, value="component_content")

        # col C (contenido) queda vacío — lo llenará el builder

    return rows_to_insert


def main():
    for label, cfg in TEMPLATES.items():
        path = cfg["path"]
        bloque = cfg["bloque"]
        try:
            wb = openpyxl.load_workbook(path)
        except PermissionError:
            print(f"  {label}: BLOQUEADO (cerrar Excel)")
            continue

        ws = wb["Secciones"]
        start, end, h3_count, _ = find_block_range(ws, bloque)
        need = TARGET_H3 - h3_count

        if need <= 0:
            print(f"  {label}: {h3_count} H3 — ya OK")
            continue

        rows_added = expand_block(ws, bloque, TARGET_H3)
        wb.save(path)
        new_total = ws.max_row
        print(f"  {label}: {h3_count}→{TARGET_H3} H3 (+{rows_added} filas, total={new_total})")

    # También expandir ImagenesComponentes para que coincidan
    expand_imagenes_componentes()


def expand_imagenes_componentes():
    """Actualiza ImagenesComponentes en las plantillas expandidas."""
    print("\n  Actualizando ImagenesComponentes...")
    # VJM: favoriteCities necesita 17 slots (antes 15-16)
    vjm_files = [
        ("VJM CIUDAD", "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_ciudad.xlsx",
         [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)]),
        ("VJM AGENCIA", "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_agencia.xlsx",
         [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)]),
        ("VJM LOCALIDAD", "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_localidad.xlsx",
         [("agencies", 2), ("favoriteCities", 17)]),
        ("VJM TIPO_AUTO", "RESULTADOS VJS/CARGUES DE CONTENIDO VJS/loadContentVjs_tipo_auto.xlsx",
         [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)]),
    ]
    for label, path, sections in vjm_files:
        try:
            wb = openpyxl.load_workbook(path)
        except PermissionError:
            print(f"    {label}: BLOQUEADO")
            continue
        ws = wb["ImagenesComponentes"]
        # Limpiar todo desde fila 3
        for r in range(3, max(ws.max_row + 1, 30)):
            for c in range(1, 13):
                ws.cell(row=r, column=c, value="")
        # Escribir nuevas filas
        row = 3
        for seccion, count in sections:
            for orden in range(1, count + 1):
                ws.cell(row=row, column=1, value=seccion)
                ws.cell(row=row, column=2, value=orden)
                row += 1
        wb.save(path)
        total = sum(c for _, c in sections)
        print(f"    {label}: {total} filas ImagenesComponentes OK")


if __name__ == "__main__":
    main()
