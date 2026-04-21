"""
fix_colors_all.py — Aplica colores EN/PT a TODOS los XLSX existentes
y renombra con ID de LP.

Post-procesa los archivos ya generados:
1. Lee col D (EN) y col E (PT) de Secciones
2. Aplica colorize_text con patrones EN/PT
3. Renombra archivo agregando IDxx al final
"""
import sys
import os
import re
import json
import sqlite3
import logging
import openpyxl
from openpyxl.cell.rich_text import CellRichText

sys.stdout.reconfigure(encoding='utf-8')
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("fix_colors")

from rich_text_formatter import colorize_text, colorize_fleet_ip

ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(ROOT, "bd_urls.db")


def extract_lp_id_from_progress(progress_file, filename):
    """Busca el lp_id en el progress JSON por el path del archivo."""
    try:
        with open(progress_file) as f:
            p = json.load(f)
        for k, v in p.items():
            if v.get("status") == "done" and v.get("path", "").endswith(filename):
                # key format: MCR_123 or VJM_123
                return int(k.split("_")[1])
    except:
        pass
    return None


def extract_lp_id_from_filename(filename):
    """Si ya tiene IDxxx en el nombre, extraer."""
    m = re.search(r'ID(\d+)', filename)
    return int(m.group(1)) if m else None


def fix_xlsx(filepath, brand, lp_id=None):
    """Aplica colores a EN/PT en un XLSX existente."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Secciones']

    fleet_block = "fleet" if brand == "mcr" else "sectionCars"

    # Encontrar filas de fleet desc/ip_usa/ip_bra para el morado
    fleet_desc_es = ""
    fleet_desc_en = ""
    fleet_desc_pt = ""

    # Primer pase: encontrar fleet desc (base para morado)
    current_block = None
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        b = (ws.cell(row, 2).value or '').strip()
        if a:
            current_block = a
        if current_block == fleet_block and b in ('Descripción H2', 'Descripción H1'):
            # Extraer texto plano de col C/D/E
            for col, lang in [(3, 'es'), (4, 'en'), (5, 'pt')]:
                cell = ws.cell(row, col)
                if isinstance(cell.value, CellRichText):
                    text = "".join(str(p) if isinstance(p, str) else p.text for p in cell.value)
                elif cell.value:
                    text = str(cell.value)
                else:
                    text = ""
                if lang == 'es': fleet_desc_es = text
                elif lang == 'en': fleet_desc_en = text
                else: fleet_desc_pt = text

    colored_count = 0

    # Segundo pase: aplicar colores a EN (col D) y PT (col E)
    current_block = None
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value or ''
        b = (ws.cell(row, 2).value or '').strip()
        if a:
            current_block = a

        for col, lang, fleet_base in [
            (3, 'es', fleet_desc_es),
            (4, 'en', fleet_desc_en),
            (5, 'pt', fleet_desc_pt),
        ]:
            cell = ws.cell(row, col)
            # Obtener texto plano (extraer de CellRichText si ya existe)
            if isinstance(cell.value, CellRichText):
                text = "".join(
                    str(p) if isinstance(p, str) else p.text
                    for p in cell.value
                )
            elif cell.value:
                text = str(cell.value)
            else:
                text = ""
            if not text or len(text) < 5:
                continue

            # Aplicar colores
            if current_block == fleet_block and b in ('IP USA', 'IP BRA'):
                rt = colorize_fleet_ip(fleet_base, text, brand, lang)
            else:
                rt = colorize_text(text, brand, lang)

            if isinstance(rt, CellRichText):
                ws.cell(row=row, column=col).value = rt
                colored_count += 1

    # Guardar
    wb.save(filepath)
    return colored_count


def main():
    # MCR
    mcr_dir = os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR")
    mcr_files = [f for f in os.listdir(mcr_dir) if "Cargue de contenido" in f and f.endswith(".xlsx")]

    log.info("=== MCR: %d archivos ===", len(mcr_files))
    mcr_progress = os.path.join(ROOT, "batch_progress.json")

    for fname in sorted(mcr_files):
        filepath = os.path.join(mcr_dir, fname)

        # Obtener LP ID
        lp_id = extract_lp_id_from_filename(fname)
        if not lp_id:
            lp_id = extract_lp_id_from_progress(mcr_progress, fname)

        try:
            colored = fix_xlsx(filepath, "mcr", lp_id)

            # Renombrar con ID si no lo tiene
            if lp_id and f"ID{lp_id}" not in fname:
                new_name = fname.replace(".xlsx", f" ID{lp_id}.xlsx")
                new_path = os.path.join(mcr_dir, new_name)
                os.rename(filepath, new_path)
                log.info("  MCR %s -> +%d colores EN/PT, renombrado ID%d", fname[:40], colored, lp_id)
            else:
                log.info("  MCR %s -> +%d colores EN/PT", fname[:40], colored)
        except Exception as e:
            log.error("  MCR %s ERROR: %s", fname[:40], e)

    # VJM
    vjm_dir = os.path.join(ROOT, "RESULTADOS VJS", "CARGUES DE CONTENIDO VJS")
    vjm_files = [f for f in os.listdir(vjm_dir) if "Cargue de contenido" in f and f.endswith(".xlsx")]

    log.info("\n=== VJM: %d archivos ===", len(vjm_files))
    vjm_progress = os.path.join(ROOT, "batch_progress_vjm.json")

    for fname in sorted(vjm_files):
        filepath = os.path.join(vjm_dir, fname)

        lp_id = extract_lp_id_from_filename(fname)
        if not lp_id:
            lp_id = extract_lp_id_from_progress(vjm_progress, fname)

        try:
            colored = fix_xlsx(filepath, "vjm", lp_id)

            if lp_id and f"ID{lp_id}" not in fname:
                new_name = fname.replace(".xlsx", f" ID{lp_id}.xlsx")
                new_path = os.path.join(vjm_dir, new_name)
                os.rename(filepath, new_path)
                log.info("  VJM %s -> +%d colores EN/PT, renombrado ID%d", fname[:40], colored, lp_id)
            else:
                log.info("  VJM %s -> +%d colores EN/PT", fname[:40], colored)
        except Exception as e:
            log.error("  VJM %s ERROR: %s", fname[:40], e)

    log.info("\nFIN")


if __name__ == "__main__":
    main()
