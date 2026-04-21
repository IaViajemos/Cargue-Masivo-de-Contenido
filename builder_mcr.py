"""
builder_mcr.py — Construye XLSX completo de cargue de contenido MCR.

Genera 1 archivo XLSX (17 hojas) por cada LP MCR, listo para subir.
Usa loadContentMcr.xlsx como template base y sobreescribe las hojas editables.
"""
import os
import re
import json
import sqlite3
import logging
import argparse
from datetime import datetime
from copy import copy

import openpyxl

from builder_landing_page import (
    construir_fila_landing_page, _get_db, obtener_urls_trinomio,
)
from rich_text_formatter import colorize_text, colorize_fleet_ip

log = logging.getLogger("builder_mcr")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(ROOT, "bd_urls.db")
TEMPLATE_PATH = os.path.join(
    ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr.xlsx"
)
OUTPUT_DIR = os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR")

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════

MCR_AGENCIAS_PRECIOS = [
    ("Budget", 8),
    ("Payless", 8),
    ("Alamo", 9),
    ("Dollar", 10),
    ("Enterprise", 10),
    ("Hertz", 10),
    ("Thrifty", 10),
    ("Avis", 11),
    ("Sixt", 12),
    ("National", 13),
]

# Slugs de agencias por idioma
AGENCY_SLUG = {
    "Alamo": "alamo", "Avis": "avis", "Budget": "budget",
    "Dollar": "dollar", "Enterprise": "enterprise", "Hertz": "hertz",
    "National": "national", "Payless": "payless", "Sixt": "sixt",
    "Thrifty": "thrifty", "Mex Rent a Car": "mex-rent-a-car",
}

# Imagenes genericas de flota (comp IDs 2-6)
FLEET_IMAGES = [
    (2, "Carros Economicos", "economicos-mcr"),
    (3, "Camionetas", "camionetas-mcr"),
    (4, "Vans", "vans-mcr"),
    (5, "Convertibles", "convertibles-mcr"),
    (6, "Carros de Lujo", "lujo-mcr"),
]

# Car type H3 titles for fleetcarrusel (posicion -> ES title)
FLEET_H3_TITLES = [
    "Carros Economicos", "Camionetas", "Vans",
    "Convertibles", "Carros de Lujo", "Carros Electricos",
]  # tipo_auto usa 6, ciudad/localidad usan 5 (trunca)

# Car type HREF slugs por idioma
FLEET_HREF = {
    "Carros Economicos": {
        "ES": "alquiler-autos-economicos", "EN": "economy-car-rental",
        "PT": "aluguel-carros-economicos",
    },
    "Camionetas": {
        "ES": "alquiler-camionetas", "EN": "suv-rental",
        "PT": "aluguel-suv",
    },
    "Vans": {
        "ES": "alquiler-vans", "EN": "van-rental",
        "PT": "aluguel-vans",
    },
    "Convertibles": {
        "ES": "alquiler-convertibles", "EN": "convertible-rental",
        "PT": "aluguel-conversivel",
    },
    "Carros de Lujo": {
        "ES": "alquiler-autos-lujo", "EN": "luxury-car-rental",
        "PT": "aluguel-carros-luxo",
    },
}

# Sentinelas de la BD
_SENTINELAS = {
    "no es localidad", "no es agencia", "no es tipo de autos",
    "no es ciudad", "no es oferta", "cat agencias", "cat tipo de autos",
}

# Mapeo tipo_lp BD -> tipo_lp generacion
TIPO_LP_MAP = {
    "AGENCIAS": "agencia",
    "LOCALIDADES": "localidad",
    "TIPOS DE AUTOS": "tipo_auto",
}

DISCLAIMER_PRECIOS = (
    "*Precios basados en los resultados entre los ultimos 12 - 24 meses."
    " Los precios pueden variar segun los dias de alquiler, el tipo de automovil "
    "y la agencia de renta de autos."
)
DISCLAIMER_FINAL = (
    "*Estos precios son sujetos a cambios  y variaran dependiendo las fechas "
    "de alquiler, agencia y tipo de vehiculo. Los precios son sugeridos por las "
    "empresas de alquiler y Miles Car Rental no asegura dichos precios. "
    "Verificar precio al momento de la reserva."
)

# ══════════════════════════════════════════════════════════════════════
# PLANTILLAS POR TIPO
# ══════════════════════════════════════════════════════════════════════

TEMPLATES_BY_TYPE = {
    "ciudad":    os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr.xlsx"),
    "agencia":   os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr agencias.xlsx"),
    "localidad": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr localidades.xlsx"),
    "tipo_auto": os.path.join(ROOT, "RESULTADOS MCR", "CARGUES DE CONTENIDO MCR", "loadContentMcr tipos autos.xlsx"),
}


# ══════════════════════════════════════════════════════════════════════
# SECCIONES MAP POR TIPO: fila Excel -> (bloque, key)
# ══════════════════════════════════════════════════════════════════════

def _faqs(start):
    """Genera mappings de 4 FAQ pairs + disclaimer final.
    start=fila de 'tit' questions. Devuelve mapa + fila disclaimer.
    """
    m = {
        start: ("questions", "tit"),
        start + 1: ("questions", "desc"),
    }
    for i in range(4):
        m[start + 2 + i*2] = ("questions", f"q_{i+1}")
        m[start + 3 + i*2] = ("questions", f"faq_{i+1}")
    m[start + 10] = ("_disclaimer_precios", None)
    return m


def _fleet_pairs(bloque, start, count):
    """Genera mappings tit/desc + count pares H3 + desc_N.
    start=fila tit. Retorna mapa con 2 + count*2 filas.
    """
    m = {
        start: (bloque, "tit"),
        start + 1: (bloque, "desc"),
    }
    for i in range(count):
        m[start + 2 + i*2] = (bloque, f"_h3_{i+1}")
        m[start + 3 + i*2] = (bloque, f"desc_{i+1}")
    return m


def _build_map_ciudad():
    """MCR CIUDAD (57 filas): qs F2-F3, fleet F8-F11, reviews F12-F13,
    rentcompanies F14-F16, questions F17-F27, fleetcarrusel F28-F39 (5 pares),
    locationscarrusel F40-F49 (4 pares), rentacar F50-F55 (2 pares), text_end F56-F57.
    """
    m = {
        2: ("quicksearch", "tit"),
        3: ("quicksearch", "desc"),
        # 4-7: benefits/agency_logs placeholders
        8: ("fleet", "tit"),
        9: ("fleet", "desc"),
        10: ("fleet", "ip_usa"),
        11: ("fleet", "ip_bra"),
        12: ("reviews", "tit"),
        13: ("reviews", "desc"),
        14: ("rentcompanies", "tit"),
        15: ("rentcompanies", "desc"),
        16: ("_disclaimer_precios", None),
    }
    m.update(_faqs(17))
    m.update(_fleet_pairs("fleetcarrusel", 28, 5))
    m.update(_fleet_pairs("locationscarrusel", 40, 17))
    m.update({
        76: ("rentacar", "tit"),
        77: ("rentacar", "desc"),
        78: ("rentacar", "tit_1"),
        79: ("rentacar", "desc_1"),
        80: ("rentacar", "tit_2"),
        81: ("rentacar", "desc_2"),
        83: ("_disclaimer_final", None),
    })
    return m


def _build_map_agencia():
    """MCR AGENCIA (71 filas): SIN fleetcarrusel.
    qs F2-3, fleet F8-11, questions F12-22, reviews F23-24,
    rentcompanies F25-27, locationscarrusel F28-63 (17 H3),
    rentacar F64-69, text_end F70, disclaimer F71.
    """
    m = {
        2: ("quicksearch", "tit"),
        3: ("quicksearch", "desc"),
        8: ("fleet", "tit"),
        9: ("fleet", "desc"),
        10: ("fleet", "ip_usa"),
        11: ("fleet", "ip_bra"),
    }
    m.update(_faqs(12))
    m.update({
        23: ("reviews", "tit"),
        24: ("reviews", "desc"),
        25: ("rentcompanies", "tit"),
        26: ("rentcompanies", "desc"),
        27: ("_disclaimer_precios", None),
    })
    m.update(_fleet_pairs("locationscarrusel", 28, 17))
    m.update({
        64: ("rentacar", "tit"),
        65: ("rentacar", "desc"),
        66: ("rentacar", "tit_1"),
        67: ("rentacar", "desc_1"),
        68: ("rentacar", "tit_2"),
        69: ("rentacar", "desc_2"),
        71: ("_disclaimer_final", None),
    })
    return m


def _build_map_localidad():
    """MCR LOCALIDAD (77 filas): SIN rentacar.
    qs F2-3, fleet F8-11, questions F12-22, reviews F23-24,
    rentcompanies F25-27, fleetcarrusel F28-39 (5 pares),
    locationscarrusel F40-75 (17 H3), disclaimer F77.
    """
    m = {
        2: ("quicksearch", "tit"),
        3: ("quicksearch", "desc"),
        8: ("fleet", "tit"),
        9: ("fleet", "desc"),
        10: ("fleet", "ip_usa"),
        11: ("fleet", "ip_bra"),
    }
    m.update(_faqs(12))
    m.update({
        23: ("reviews", "tit"),
        24: ("reviews", "desc"),
        25: ("rentcompanies", "tit"),
        26: ("rentcompanies", "desc"),
        27: ("_disclaimer_precios", None),
    })
    m.update(_fleet_pairs("fleetcarrusel", 28, 5))
    m.update(_fleet_pairs("locationscarrusel", 40, 17))
    m[77] = ("_disclaimer_final", None)
    return m


def _build_map_tipo_auto():
    """MCR TIPO_AUTO (85 filas): fleetcarrusel 6 pares, locationscarrusel 17 H3.
    qs F2-3, fleet F8-11, questions F12-22, reviews F23-24,
    locationscarrusel F25-60 (17 H3), rentcompanies F61-63,
    fleetcarrusel F64-77 (6 pares), rentacar F78-83 (2 pares),
    text_end F84, disclaimer F85.
    """
    m = {
        2: ("quicksearch", "tit"),
        3: ("quicksearch", "desc"),
        8: ("fleet", "tit"),
        9: ("fleet", "desc"),
        10: ("fleet", "ip_usa"),
        11: ("fleet", "ip_bra"),
    }
    m.update(_faqs(12))
    m.update({
        23: ("reviews", "tit"),
        24: ("reviews", "desc"),
    })
    m.update(_fleet_pairs("locationscarrusel", 25, 17))
    m.update({
        61: ("rentcompanies", "tit"),
        62: ("rentcompanies", "desc"),
        63: ("_disclaimer_precios", None),
    })
    m.update(_fleet_pairs("fleetcarrusel", 64, 6))
    m.update({
        78: ("rentacar", "tit"),
        79: ("rentacar", "desc"),
        80: ("rentacar", "tit_1"),
        81: ("rentacar", "desc_1"),
        82: ("rentacar", "tit_2"),
        83: ("rentacar", "desc_2"),
        85: ("_disclaimer_final", None),
    })
    return m


SECCIONES_MAP_MCR = {
    "ciudad":    _build_map_ciudad(),
    "agencia":   _build_map_agencia(),
    "localidad": _build_map_localidad(),
    "tipo_auto": _build_map_tipo_auto(),
}

# Back-compat: SECCIONES_MAP (default ciudad)
SECCIONES_MAP = SECCIONES_MAP_MCR["ciudad"]

# H3 rows y filas fleetcarrusel por tipo (para fill_secciones)
FLEETCARRUSEL_H3_ROWS_BY_TYPE = {
    "ciudad":    [30, 32, 34, 36, 38],
    "agencia":   [],  # sin fleetcarrusel
    "localidad": [30, 32, 34, 36, 38],
    "tipo_auto": [40, 42, 44, 46, 48, 50],
}

LOCATIONSCARRUSEL_H3_ROWS_BY_TYPE = {
    "ciudad":    list(range(42, 76, 2)),  # F42,44,...,74 = 17 H3
    "agencia":   list(range(30, 64, 2)),  # F30,32,...,62 = 17 H3
    "localidad": list(range(42, 76, 2)),  # F42,44,...,74 = 17 H3
    "tipo_auto": list(range(27, 61, 2)),  # F27,29,...,59 = 17 H3
}


# ══════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def _es_real(val):
    """True si el valor NO es sentinela ni vacio."""
    if not val:
        return False
    return val.strip().lower() not in _SENTINELAS


def _slugify(text):
    """Convierte texto a slug URL: lowercase, sin acentos, guiones."""
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")


def extract_country_slug(url_path):
    """Extrae el slug del pais del url_path ES."""
    parts = url_path.strip("/").split("/")
    if len(parts) >= 2:
        seg = parts[1]
        skip = {"agencias", "localidades", "info", "ofertas", "tipos-autos",
                "tipos-carros", "cat-agencias", "cat-tipo-de-autos"}
        if seg not in skip and not seg.startswith("alquiler-") and not seg.startswith("aluguel-"):
            return seg
    return ""


def extract_dominio_suffix(dominio):
    """Extrae el sufijo del dominio: '.com', 'tampa.com', '.eu', etc."""
    m = re.search(r"milescarrental(\w*)\.(com|eu|de|fr|it|pt|co\.uk|com\.mx|com\.br|com\.es)", dominio)
    if m:
        city = m.group(1)
        tld = m.group(2)
        return f"{city}.{tld}" if city else f".{tld}"
    return ".com"


# ══════════════════════════════════════════════════════════════════════
# CONTEXTO LP
# ══════════════════════════════════════════════════════════════════════

def build_ctx(db, lp_id):
    """Construye dict de contexto normalizado para una LP."""
    row = db.execute("SELECT * FROM landing_pages WHERE id = ?", (lp_id,)).fetchone()
    if not row:
        raise ValueError(f"LP {lp_id} no encontrada")

    lp = dict(row)
    trinomio = obtener_urls_trinomio(db, lp_id)

    ciudad = lp["ciudad"] if _es_real(lp["ciudad"]) else ""
    localidad = lp["localidad"] if _es_real(lp["localidad"]) else ""
    agencia = lp["agencia"] if _es_real(lp["agencia"]) else ""
    car_category = lp["car_category"] if _es_real(lp["car_category"]) else ""
    dominio = lp["dominio"] or "https://www.milescarrental.com/"
    tipo_lp_bd = lp["tipo_lp"]

    url_es = trinomio.get("ES", {}).get("url_path", "")
    url_en = trinomio.get("EN", {}).get("url_path", "")
    url_pt = trinomio.get("PT", {}).get("url_path", "")
    country_slug = extract_country_slug(url_es)
    dominio_suffix = extract_dominio_suffix(dominio)

    # Determinar tipo LP para generacion
    # tipo_lp_bd='LOCALIDADES' sin localidad real es una CIUDAD pura
    if tipo_lp_bd == "LOCALIDADES" and not localidad:
        tipo_gen = "ciudad"
    elif tipo_lp_bd in TIPO_LP_MAP:
        tipo_gen = TIPO_LP_MAP[tipo_lp_bd]
    else:
        tipo_gen = "ciudad"

    # Nombre para imagenes
    if tipo_lp_bd == "AGENCIAS" and agencia:
        if ciudad:
            img_prefix = _slugify(f"{agencia} {ciudad}")
        else:
            img_prefix = f"{_slugify(agencia)}-mundo"
    elif tipo_lp_bd == "TIPOS DE AUTOS" and car_category:
        if ciudad:
            img_prefix = _slugify(f"{car_category} {ciudad}")
        else:
            img_prefix = f"{_slugify(car_category)}-mundo"
    elif ciudad:
        img_prefix = _slugify(ciudad)
    else:
        img_prefix = "default"

    # Nombre representativo para titulos
    if tipo_lp_bd == "AGENCIAS" and agencia:
        nombre_lp = f"{agencia} {ciudad}" if ciudad else agencia
    elif tipo_lp_bd == "TIPOS DE AUTOS" and car_category:
        nombre_lp = f"{car_category} {ciudad}" if ciudad else car_category
    elif tipo_lp_bd == "LOCALIDADES" and localidad:
        nombre_lp = localidad
    else:
        nombre_lp = ciudad

    return {
        "lp_id": lp_id,
        "lp": lp,
        "tipo_lp_bd": tipo_lp_bd,
        "tipo_gen": tipo_gen,
        "ciudad": ciudad,
        "localidad": localidad,
        "agencia": agencia,
        "car_category": car_category,
        "dominio": dominio,
        "dominio_suffix": dominio_suffix,
        "country_slug": country_slug,
        "img_prefix": img_prefix,
        "nombre_lp": nombre_lp,
        "urls": {
            "ES": url_es,
            "EN": url_en,
            "PT": url_pt,
        },
    }


# ══════════════════════════════════════════════════════════════════════
# FILL FUNCTIONS
# ══════════════════════════════════════════════════════════════════════

def fill_landing_page(ws, ctx, db):
    """Hoja LandingPage — fila 3 con datos de BD."""
    fila = construir_fila_landing_page(db, ctx["lp_id"])
    col_map = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7,
               "H": 8, "I": 9, "J": 10, "K": 11, "L": 12, "M": 13}
    for col_letter, val in fila.items():
        ws.cell(row=3, column=col_map[col_letter], value=val)


def fill_secciones(ws, ctx, content, tipo_gen=None,
                   content_en=None, content_pt=None):
    """Hoja Secciones - escribe ES (col C), EN (col D), PT (col E).

    content: dict bloques ES
    content_en: dict bloques EN (traducido)
    content_pt: dict bloques PT (traducido)
    """
    tipo_gen = tipo_gen or ctx.get("tipo_gen", "ciudad")
    mapa = SECCIONES_MAP_MCR.get(tipo_gen, SECCIONES_MAP_MCR["ciudad"])
    content_en = content_en or {}
    content_pt = content_pt or {}

    for row, (bloque, key) in mapa.items():
        if bloque == "_disclaimer_precios":
            ws.cell(row=row, column=3, value=DISCLAIMER_PRECIOS)
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
            continue
        if bloque == "_disclaimer_final":
            ws.cell(row=row, column=3, value=DISCLAIMER_FINAL)
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
            continue

        if key and key.startswith("_h3_"):
            continue  # Se manejan aparte

        bloque_data = content.get(bloque, {})
        val_es = bloque_data.get(key, "") if (bloque_data and key) else ""
        bloque_en = content_en.get(bloque, {})
        bloque_pt = content_pt.get(bloque, {})
        val_en = bloque_en.get(key, "") if (bloque_en and key) else ""
        val_pt = bloque_pt.get(key, "") if (bloque_pt and key) else ""

        # Aplicar rich text (colores inline) a ES, EN, PT
        desc_base_es = bloque_data.get("desc", "") if bloque == "fleet" else ""
        desc_base_en = bloque_en.get("desc", "") if bloque == "fleet" else ""
        desc_base_pt = bloque_pt.get("desc", "") if bloque == "fleet" else ""

        for col, val, lang, base in [
            (3, val_es, "es", desc_base_es),
            (4, val_en, "en", desc_base_en),
            (5, val_pt, "pt", desc_base_pt),
        ]:
            if val and bloque == "fleet" and key in ("ip_usa", "ip_bra"):
                ws.cell(row=row, column=col).value = colorize_fleet_ip(
                    base, val, "mcr", lang)
            elif val:
                ws.cell(row=row, column=col).value = colorize_text(val, "mcr", lang)
            else:
                ws.cell(row=row, column=col, value="")

    # fleetcarrusel H3 titles (filas segun tipo)
    FLEET_H3_EN = [
        "Economy Cars", "SUVs", "Vans",
        "Convertibles", "Luxury Cars", "Electric Cars",
    ]
    FLEET_H3_PT = [
        "Carros Econômicos", "Caminhonetes", "Vans",
        "Conversíveis", "Carros de Luxo", "Carros Elétricos",
    ]
    fc = content.get("fleetcarrusel", {})
    fc_rows = FLEETCARRUSEL_H3_ROWS_BY_TYPE.get(tipo_gen, [])
    for i, row in enumerate(fc_rows):
        es = FLEET_H3_TITLES[i] if i < len(FLEET_H3_TITLES) else ""
        en = FLEET_H3_EN[i] if i < len(FLEET_H3_EN) else ""
        pt = FLEET_H3_PT[i] if i < len(FLEET_H3_PT) else ""
        ws.cell(row=row, column=3, value=es if fc else "")
        ws.cell(row=row, column=4, value=en if fc else "")
        ws.cell(row=row, column=5, value=pt if fc else "")

    # locationscarrusel H3 titles + traducciones
    lc = content.get("locationscarrusel", {})
    lc_en = content_en.get("locationscarrusel", {})
    lc_pt = content_pt.get("locationscarrusel", {})
    lc_rows = LOCATIONSCARRUSEL_H3_ROWS_BY_TYPE.get(tipo_gen, [])
    for i, row in enumerate(lc_rows):
        tit_es = lc.get(f"tit_{i+1}", "") if lc else ""
        tit_en = lc_en.get(f"tit_{i+1}", "") if lc_en else ""
        tit_pt = lc_pt.get(f"tit_{i+1}", "") if lc_pt else ""
        ws.cell(row=row, column=3, value=tit_es or "")
        ws.cell(row=row, column=4, value=tit_en or "")
        ws.cell(row=row, column=5, value=tit_pt or "")


def fill_boton_componentes(ws, ctx):
    """Hoja BotonComponentes — estatico, ya en template."""
    # Validar/sobreescribir para asegurar datos correctos
    comp_ids = list(range(2, 14))  # 2-13
    for i, cid in enumerate(comp_ids):
        row = 3 + i
        ws.cell(row=row, column=1, value=cid)
        ws.cell(row=row, column=2, value="Ver Ofertas")
        ws.cell(row=row, column=3, value="See Offers")
        ws.cell(row=row, column=4, value="Ver Ofertas")


def fill_precios_agencias(ws, ctx):
    """Hoja PreciosAgencias — agencias + precios + HREFs trilingual."""
    tipo = ctx["tipo_lp_bd"]
    country = ctx["country_slug"]
    ciudad = ctx["ciudad"]
    localidad = ctx["localidad"]
    es_ciudad_pura = (tipo == "LOCALIDADES" and not localidad and ciudad)

    for i, (agencia, precio) in enumerate(MCR_AGENCIAS_PRECIOS):
        row = 3 + i
        slug = AGENCY_SLUG.get(agencia, _slugify(agencia))

        # HREFs segun tipo LP
        if tipo == "AGENCIAS":
            href_es = f"/es/alquiler-autos-{slug}"
            href_en = f"/en/{slug}-car-rental"
            href_pt = f"/pt/aluguel-carros-{slug}"
        elif es_ciudad_pura and country:
            # Ciudad pura (LOCALIDADES sin localidad): /es/{country}/alquiler-autos-{slug}
            href_es = f"/es/{country}/alquiler-autos-{slug}"
            href_en = f"/en/{country}/{slug}-car-rental"
            href_pt = f"/pt/{country}/aluguel-carros-{slug}"
        elif tipo in ("LOCALIDADES", "TIPOS DE AUTOS"):
            href_es = f"/es/agencias/alquiler-autos-{slug}"
            href_en = f"/en/agencies/{slug}-car-rental"
            href_pt = f"/pt/locadoras/aluguel-carros-{slug}"
        else:
            # Ciudad: /es/{country}/alquiler-autos-{slug}
            if country:
                href_es = f"/es/{country}/alquiler-autos-{slug}"
                href_en = f"/en/{country}/{slug}-car-rental"
                href_pt = f"/pt/{country}/aluguel-carros-{slug}"
            else:
                href_es = f"/es/alquiler-autos-{slug}"
                href_en = f"/en/{slug}-car-rental"
                href_pt = f"/pt/aluguel-carros-{slug}"

        # Titulos
        if ciudad and tipo != "AGENCIAS":
            tit_es = f"{agencia} Rent a Car en {ciudad}"
            tit_en = f"{agencia} Rent a Car in {ciudad}"
            tit_pt = f"{agencia} Rent a Car em {ciudad}"
        else:
            tit_es = f"{agencia} Rent a Car"
            tit_en = f"{agencia} Rent a Car"
            tit_pt = f"{agencia} Rent a Car"

        ws.cell(row=row, column=1, value=1)       # COMPONENTE ID
        ws.cell(row=row, column=2, value=agencia)  # AGENCIA
        ws.cell(row=row, column=3, value=precio)   # PRECIO
        ws.cell(row=row, column=4, value=href_es)
        ws.cell(row=row, column=5, value=tit_es)
        ws.cell(row=row, column=6, value=href_en)
        ws.cell(row=row, column=7, value=tit_en)
        ws.cell(row=row, column=8, value=href_pt)
        ws.cell(row=row, column=9, value=tit_pt)


def fill_imagenes_secciones(ws, ctx):
    """Hoja ImagenesSecciones — hero (sec 1) + opiniones (sec 3)."""
    prefix = ctx["img_prefix"]
    ciudad = ctx["ciudad"]
    nombre = ctx["nombre_lp"]

    # Row 3: Section 1 (quicksearch hero)
    jpg = f"images/ContentLp/{prefix}-mcr.jpg"
    webp = f"images/ContentLp/{prefix}-mcr.webp"
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value=jpg)
    ws.cell(row=3, column=3, value=webp)
    ws.cell(row=3, column=4, value=f"Renta de carros en {nombre}")
    ws.cell(row=3, column=5, value=f"Alquiler de autos en {nombre}")
    ws.cell(row=3, column=6, value=None)  # HREF
    ws.cell(row=3, column=7, value=f"Car rentals in {nombre}")
    ws.cell(row=3, column=8, value=f"Car rentals in {nombre}")
    ws.cell(row=3, column=9, value=None)
    ws.cell(row=3, column=10, value=f"Aluguel de carros em {nombre}")
    ws.cell(row=3, column=11, value=f"Locacao de veiculos em {nombre}")
    ws.cell(row=3, column=12, value=None)

    # Row 4: Section 3 (reviews)
    jpg_op = f"images/ContentLp/{prefix}-mcr-opiniones.jpg"
    webp_op = f"images/ContentLp/{prefix}-mcr-opiniones.webp"
    ws.cell(row=4, column=1, value=3)
    ws.cell(row=4, column=2, value=jpg_op)
    ws.cell(row=4, column=3, value=webp_op)
    ws.cell(row=4, column=4, value=f"Opiniones sobre alquiler de autos en {nombre}")
    ws.cell(row=4, column=5, value=f"Resenas sobre renta de carros en {nombre}")
    ws.cell(row=4, column=6, value=None)
    ws.cell(row=4, column=7, value=f"Reviews about car rentals in {nombre}")
    ws.cell(row=4, column=8, value=f"Customer reviews about car rentals in {nombre}")
    ws.cell(row=4, column=9, value=None)
    ws.cell(row=4, column=10, value=f"Opinioes sobre aluguel de carros em {nombre}")
    ws.cell(row=4, column=11, value=f"Comentarios sobre locacao de veiculos em {nombre}")
    ws.cell(row=4, column=12, value=None)


def fill_imagenes_componentes(ws, ctx):
    """Hoja ImagenesComponentes — flota generica (comp 2-6) + cluster (comp 7-10)."""
    nombre = ctx["nombre_lp"]
    tipo = ctx["tipo_lp_bd"]
    country = ctx["country_slug"]

    # Flota generica (siempre las mismas 5 imagenes)
    for i, (comp_id, car_name, img_base) in enumerate(FLEET_IMAGES):
        row = 3 + i
        jpg = f"images/ContentLp/{img_base}.jpg"
        webp = f"images/ContentLp/{img_base}.webp"
        href_data = FLEET_HREF.get(car_name, {})

        # HREFs para tipos de flota
        if country and tipo not in ("AGENCIAS", "LOCALIDADES", "TIPOS DE AUTOS"):
            href_es = f"/es/{country}/{href_data.get('ES', '')}"
            href_en = f"/en/{country}/{href_data.get('EN', '')}"
            href_pt = f"/pt/{country}/{href_data.get('PT', '')}"
        else:
            href_es = f"/es/{href_data.get('ES', '')}"
            href_en = f"/en/{href_data.get('EN', '')}"
            href_pt = f"/pt/{href_data.get('PT', '')}"

        ws.cell(row=row, column=1, value=comp_id)
        ws.cell(row=row, column=2, value=jpg)
        ws.cell(row=row, column=3, value=webp)
        ws.cell(row=row, column=4, value=f"Renta de autos en {nombre}, al mejor precio")
        ws.cell(row=row, column=5, value=f"Alquiler de carros baratos en {nombre}")
        ws.cell(row=row, column=6, value=href_es)
        ws.cell(row=row, column=7, value=f"Car rentals in {nombre}, at the best price")
        ws.cell(row=row, column=8, value=f"Affordable car rentals in {nombre}")
        ws.cell(row=row, column=9, value=href_en)
        ws.cell(row=row, column=10, value=f"Aluguel de carros em {nombre}, ao melhor preco")
        ws.cell(row=row, column=11, value=f"Locacao de veiculos economicos em {nombre}")
        ws.cell(row=row, column=12, value=href_pt)

    # Cluster locations (comp 7-10): dejar vacio por ahora
    # TODO: poblar con ciudades cercanas cuando se defina la logica


def fill_galerias_componentes(ws, ctx):
    """Hoja GaleriasComponentes — 1 fila."""
    slug = ctx["img_prefix"]
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=2, value=11)  # COMPONENTE ID
    ws.cell(row=3, column=3, value=slug)


def fill_imagenes_galerias(ws, ctx):
    """Hoja ImagenesGalerias — 3 imagenes de blog con alt texts variados."""
    prefix = ctx["img_prefix"]
    nombre = ctx["nombre_lp"]

    alt_es = [
        f"Atracciones turisticas en {nombre}",
        f"Actividades para disfrutar en {nombre}",
        f"Lugares imperdibles en {nombre}",
    ]
    alt_en = [
        f"Tourist attractions in {nombre}",
        f"Activities to enjoy in {nombre}",
        f"Must-see places in {nombre}",
    ]
    alt_pt = [
        f"Atracoes turisticas em {nombre}",
        f"Atividades para aproveitar em {nombre}",
        f"Lugares imperdíveis em {nombre}",
    ]
    tit_es = [
        f"Mejores lugares para visitar en {nombre}",
        f"Planes y actividades en {nombre}",
        f"Explora lo mejor de {nombre}",
    ]

    for pos in range(1, 4):
        row = 2 + pos
        i = pos - 1
        jpg = f"images/ContentLp/{prefix}-mcr-blog-{pos}.jpg"
        webp = f"images/ContentLp/{prefix}-mcr-blog-{pos}.webp"

        ws.cell(row=row, column=1, value=1)   # GALERIA ID
        ws.cell(row=row, column=2, value=jpg)
        ws.cell(row=row, column=3, value=webp)
        ws.cell(row=row, column=4, value=alt_es[i])
        ws.cell(row=row, column=5, value=tit_es[i])
        ws.cell(row=row, column=6, value=None)
        ws.cell(row=row, column=7, value=alt_en[i])
        ws.cell(row=row, column=8, value=f"Things to do in {nombre}")
        ws.cell(row=row, column=9, value=None)
        ws.cell(row=row, column=10, value=alt_pt[i])
        ws.cell(row=row, column=11, value=f"Pontos turisticos em {nombre}")
        ws.cell(row=row, column=12, value=None)
        ws.cell(row=row, column=13, value=pos)


def fill_categorias_flota(ws, ctx):
    """Hoja CategoriasFlotaLandingPage — default."""
    ws.cell(row=3, column=1, value="Pequenos y Medianos")
    for c in range(2, 8):
        ws.cell(row=3, column=c, value="")


def fill_tablero_control(ws, ctx, output_filename):
    """Hoja Tablero de Control."""
    ws.cell(row=1, column=1, value="Nombre plantilla")
    ws.cell(row=1, column=2, value="Cargue de contenido")
    ws.cell(row=2, column=1, value="Dominio")
    ws.cell(row=2, column=2, value="MCR")
    ws.cell(row=3, column=1, value="LP")
    ws.cell(row=3, column=2, value=ctx["nombre_lp"])
    ws.cell(row=4, column=1, value="Ciudad")
    ws.cell(row=4, column=2, value=ctx["dominio_suffix"])
    ws.cell(row=5, column=1, value="Nombre Archivo")
    ws.cell(row=5, column=2, value=output_filename)


# ══════════════════════════════════════════════════════════════════════
# ORQUESTADOR
# ══════════════════════════════════════════════════════════════════════

def build_output_filename(ctx):
    """Genera nombre de archivo de salida con ID de LP."""
    fecha = datetime.now().strftime("%d-%m-%Y")
    nombre = ctx["nombre_lp"] or "SinNombre"
    sufijo = ctx["dominio_suffix"]
    lp_id = ctx["lp_id"]
    return f"{fecha} Cargue de contenido MCR {nombre} {sufijo} ID{lp_id}.xlsx"


def _template_for_tipo(tipo_gen):
    return TEMPLATES_BY_TYPE.get(tipo_gen, TEMPLATES_BY_TYPE["ciudad"])


def build_single(db, lp_id, content=None, content_en=None, content_pt=None,
                 template_path=None, output_dir=OUTPUT_DIR):
    """Construye 1 XLSX MCR completo con traducciones EN/PT."""
    ctx = build_ctx(db, lp_id)
    tipo_gen = ctx["tipo_gen"]
    tmpl = template_path or _template_for_tipo(tipo_gen)

    log.info("Construyendo MCR LP %d: %s (%s) plantilla=%s",
             lp_id, ctx["nombre_lp"], ctx["tipo_lp_bd"], os.path.basename(tmpl))

    wb = openpyxl.load_workbook(tmpl)
    output_name = build_output_filename(ctx)

    fill_landing_page(wb["LandingPage"], ctx, db)
    fill_precios_agencias(wb["PreciosAgencias"], ctx)
    fill_imagenes_secciones(wb["ImagenesSecciones"], ctx)
    fill_imagenes_galerias(wb["ImagenesGalerias"], ctx)
    fill_categorias_flota(wb["CategoriasFlotaLandingPage"], ctx)
    fill_tablero_control(wb["Tablero de Control"], ctx, output_name)

    if content:
        fill_secciones(wb["Secciones"], ctx, content, tipo_gen=tipo_gen,
                       content_en=content_en, content_pt=content_pt)

    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)
    log.info("Guardado: %s", output_path)
    return output_path


def build_from_json(json_path, template_path=TEMPLATE_PATH, output_dir=OUTPUT_DIR):
    """Construye XLSX para cada LP MCR en el JSON de pruebas."""
    with open(json_path, "r", encoding="utf-8") as f:
        all_lps = json.load(f)

    db = _get_db()
    results = []

    for lp_data in all_lps:
        if lp_data.get("marca") != "MCR":
            continue

        ciudad = lp_data["ciudad"]
        tipo = lp_data["tipo_lp"]
        log.info("Procesando MCR %s: %s", tipo, ciudad)

        # Buscar LP en BD por ciudad y tipo
        row = db.execute(
            "SELECT id FROM landing_pages WHERE proyecto='MCR' AND ciudad LIKE ? LIMIT 1",
            (f"%{ciudad}%",)
        ).fetchone()

        if not row:
            log.warning("LP no encontrada en BD para: %s", ciudad)
            continue

        # Convertir bloques de lista a dict
        content = {}
        for bloque in lp_data.get("bloques", []):
            content[bloque["nombre"]] = bloque["contenido"]

        path = build_single(db, row["id"], content=content,
                            template_path=template_path, output_dir=output_dir)
        results.append(path)

    db.close()
    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Builder MCR XLSX")
    parser.add_argument("--json", help="Ruta al JSON de pruebas")
    parser.add_argument("--lp-id", type=int, help="ID de LP especifica")
    parser.add_argument("--output-dir", default=OUTPUT_DIR)
    args = parser.parse_args()

    if args.json:
        paths = build_from_json(args.json, output_dir=args.output_dir)
        print(f"\nGenerados {len(paths)} XLSX:")
        for p in paths:
            print(f"  {os.path.basename(p)}")
    elif args.lp_id:
        db = _get_db()
        path = build_single(db, args.lp_id, output_dir=args.output_dir)
        db.close()
        print(f"Generado: {path}")
    else:
        # Demo: generar desde JSON de pruebas
        json_path = os.path.join(ROOT, "pruebas_8lp_contenido_completo.json")
        if os.path.exists(json_path):
            paths = build_from_json(json_path, output_dir=args.output_dir)
            print(f"\nGenerados {len(paths)} XLSX:")
            for p in paths:
                print(f"  {os.path.basename(p)}")
        else:
            print("Uso: python builder_mcr.py --json pruebas.json")
            print("     python builder_mcr.py --lp-id 123")
