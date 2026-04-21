import openpyxl
from collections import Counter

path = r"C:\Users\Miguel Martinez SSD\OneDrive - BROWSER TRAVEL SOLUTIONS S.A.S VIAJEMOS\Documentos\PROYECTOS\CARGUE MASIVO VJM Y MCR\RESULTADOS VJS\URLS Y BASE VJS\BD carros nuevas URL VJS - V.31.03.26.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

data = []
for row in range(2, ws.max_row + 1):
    data.append({
        'row': row,
        'dominio': ws.cell(row, 4).value,
        'negocio': ws.cell(row, 3).value,
        'pais': ws.cell(row, 7).value,
        'ciudad': ws.cell(row, 8).value,
        'sub_localidad': ws.cell(row, 9).value,
        'agencias': ws.cell(row, 6).value,
        'car_category': ws.cell(row, 12).value,
        'agencia': ws.cell(row, 13).value,
        'idioma': ws.cell(row, 10).value,
        'cant_idiomas': ws.cell(row, 11).value,
        'ofertas': ws.cell(row, 14).value,
        'url': ws.cell(row, 17).value,
    })

# Show the 11 groups of 6 in detail
print("=== 6-ROW GROUPS (AGENCIAS+key) — DETAIL ===")
with_cat = {}
for d in data:
    key = (d['dominio'], d['pais'], d['ciudad'], d['sub_localidad'], d['agencia'], d['car_category'], d['ofertas'], d['agencias'])
    if key not in with_cat:
        with_cat[key] = []
    with_cat[key].append(d)

for key, rows in with_cat.items():
    if len(rows) == 6:
        print(f"\n  KEY: {key}")
        for r in rows:
            print(f"    row {r['row']}: {r['idioma']} | url={r['url']}")

# Show spelling mismatches: rows where positional group != field-based group
print("\n\n=== SPELLING MISMATCHES (EN vs ES/PT city names) ===")
# Check some known mismatches
sample_checks = [
    (926, 928),  # Bahia Blanca
    (971, 973),  # Florianopolis
    (977, 979),  # Maceio
]
for start, end in sample_checks:
    if end <= len(data):
        print(f"\n  Rows {start}-{end}:")
        for r in range(start, end + 1):
            idx = r - 2
            if 0 <= idx < len(data):
                d = data[idx]
                print(f"    row {r}: {d['idioma']} | ciudad={repr(d['ciudad'])} | url={d['url']}")

# FINAL ANALYSIS: Positional grouping as the correct strategy
print("\n\n========================================")
print("=== FINAL: POSITIONAL GROUPING SUMMARY ===")
print("========================================")

groups_pos = []
i = 0
while i < len(data):
    group = [data[i]]
    j = i + 1
    # Use cant_idiomas to know expected group size
    expected = data[i].get('cant_idiomas', 3)
    if not isinstance(expected, int):
        expected = 3  # default
    while j < len(data) and len(group) < expected:
        nxt = data[j]
        # Check language is different from all in group so far
        if nxt['idioma'] not in [g['idioma'] for g in group]:
            group.append(nxt)
            j += 1
        else:
            break
    groups_pos.append(group)
    i = j

print(f"Positional groups: {len(groups_pos)}")
sz = Counter(len(g) for g in groups_pos)
for s, c in sorted(sz.items()):
    print(f"  {s}-lang: {c} groups")
print(f"  Total rows: {sum(len(g) for g in groups_pos)}")

# Show which AGENCIAS categories have which group counts
print("\n=== GROUPS BY AGENCIAS CATEGORY ===")
cat_groups = {}
for g in groups_pos:
    cat = g[0]['agencias']
    if cat not in cat_groups:
        cat_groups[cat] = 0
    cat_groups[cat] += 1
for cat, cnt in sorted(cat_groups.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {cnt} landing pages")

# Show what the old grouping key might have been to get 369
# Try: (dominio, negocio, pais, ciudad, sub_localidad, agencias, car_category) but ignoring idioma and counting unique
print("\n\n=== REPRODUCE 369 INVESTIGATION ===")
# Maybe old code grouped all rows (not just unique combos minus idioma)?
# Or maybe it used a different subset?
# Let's try: only AGENCIAS rows
agencias_only = [d for d in data if d['agencias'] == 'AGENCIAS']
print(f"Rows where AGENCIAS='AGENCIAS': {len(agencias_only)}")
ag_groups = Counter()
for d in agencias_only:
    key = (d['dominio'], d['pais'], d['ciudad'], d['sub_localidad'], d['car_category'])
    ag_groups[key] += 1
print(f"Unique groups (AGENCIAS only): {len(ag_groups)}")
ag_size = Counter(ag_groups.values())
for s, c in sorted(ag_size.items()):
    print(f"  {s} rows: {c} groups")

# Try: (pais, ciudad, sub_localidad) only
geo_groups = Counter()
for d in data:
    key = (d['pais'], d['ciudad'], d['sub_localidad'])
    geo_groups[key] += 1
print(f"\nGeo groups (pais+ciudad+sub_localidad): {len(geo_groups)}")

# try (dominio, pais, ciudad, sub_localidad)
geo2 = Counter()
for d in data:
    key = (d['dominio'], d['pais'], d['ciudad'], d['sub_localidad'])
    geo2[key] += 1
print(f"Dominio+geo groups: {len(geo2)}")

# Check if maybe they divided by idioma count somehow
# 1107 ES rows / 3 = 369 exactly!
print(f"\nES rows: {sum(1 for d in data if d['idioma'] == 'ES')}")
print(f"ES rows / 3 = {sum(1 for d in data if d['idioma'] == 'ES') / 3}")

# Or: unique ES URLs
es_urls = set(d['url'] for d in data if d['idioma'] == 'ES')
print(f"Unique ES URLs: {len(es_urls)}")

# Check how many unique combos when filtering only ES
es_data = [d for d in data if d['idioma'] == 'ES']
es_combos = Counter()
for d in es_data:
    key = (d['dominio'], d['negocio'], d['pais'], d['ciudad'], d['sub_localidad'], d['agencias'], d['car_category'])
    es_combos[key] += 1
print(f"ES only with old 7-field key: {len(es_combos)} unique")

# Maybe: old key without car_category and agencia
reduced = Counter()
for d in data:
    key = (d['dominio'], d['negocio'], d['pais'], d['ciudad'], d['sub_localidad'], d['agencias'])
    reduced[key] += 1
print(f"\nReduced key (no car_category): {len(reduced)} groups")
