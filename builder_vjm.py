"""
builder_vjm.py - Construye XLSX completo de cargue de contenido Viajemos.

Usa plantillas específicas POR TIPO de LP:
  - loadContentVjs_ciudad.xlsx (70 filas Secciones)
  - loadContentVjs_agencia.xlsx (68 filas)
  - loadContentVjs_localidad.xlsx (55 filas, sin carRental)
  - loadContentVjs_tipo_auto.xlsx (68 filas)

Hojas editables (6): LandingPage, Secciones, ImagenesSecciones,
PreciosAgencias, ImagenesComponentes, categoriaFlotaVehiculos.
Hojas catalogo read-only: Companies, Agencias, FormatoContenido, FuenteTextos.
"""
import os
import re
import json
import logging
from datetime import datetime

import openpyxl

from builder_landing_page import (
    _get_db, obtener_urls_trinomio,
    generar_meta_titulo, generar_meta_descripcion,
)
from rich_text_formatter import colorize_text, colorize_fleet_ip

log = logging.getLogger("builder_vjm")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

ROOT = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(ROOT, "bd_urls.db")
TEMPLATES_DIR = os.path.join(ROOT, "RESULTADOS VJS", "CARGUES DE CONTENIDO VJS")
OUTPUT_DIR = TEMPLATES_DIR

TEMPLATES_BY_TYPE = {
    "ciudad":    os.path.join(TEMPLATES_DIR, "loadContentVjs_ciudad.xlsx"),
    "agencia":   os.path.join(TEMPLATES_DIR, "loadContentVjs_agencia.xlsx"),
    "localidad": os.path.join(TEMPLATES_DIR, "loadContentVjs_localidad.xlsx"),
    "tipo_auto": os.path.join(TEMPLATES_DIR, "loadContentVjs_tipo_auto.xlsx"),
}

# ══════════════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════════════

VJM_HIGH_PRICE = 182.02
VJM_LOW_PRICE = 9

VJM_AGENCIAS = [
    ("AL_Alamo", "Alamo", 9),
    ("ZI_Avis", "Avis", 11),
    ("ZD_Budget", "Budget", 8),
    ("ZR_Dollar", "Dollar", 10),
    ("ET_Enterprise", "Enterprise", 10),
    ("ZE_Hertz", "Hertz", 10),
    ("ZL_National", "National", 13),
    ("ZA_Payless", "Payless", 8),
    ("SX_Sixt", "Sixt", 12),
    ("ZT_Thrifty", "Thrifty", 10),
]

AGENCY_SLUG = {
    "Alamo": "alamo", "Avis": "avis", "Budget": "budget",
    "Dollar": "dollar", "Enterprise": "enterprise", "Hertz": "hertz",
    "National": "national", "Payless": "payless", "Sixt": "sixt",
    "Thrifty": "thrifty",
}

DOMINIO_BD_TO_SLUG = {
    ".COM VJS": "viajemos",
    "ARGENTINA VJS": "viajemos_argentina",
    "BOLIVIA VJS": "viajemos_bolivia",
    "BRASIL VJS": "viajemos_brasil",
    "CANADÁ VJS": "viajemos_canada",
    "CHILE VJS": "viajemos_chile",
    "COLOMBIA VJS": "viajemos_colombia",
    "COSTA RICA VJS": "viajemos_costa_rica",
    "ECUADOR VJS": "viajemos_ecuador",
    "ESPAÑA VJS": "viajemos_espana",
    "GUATEMALA VJS": "viajemos_guatemala",
    "HONDURAS VJS": "viajemos_honduras",
    "MÉXICO VJS": "viajemos_mexico",
    "NICARAGUA VJS": "viajemos_nicaragua",
    "PANAMÁ VJS": "viajemos_panama",
    "PARAGUAY VJS": "viajemos_paraguay",
    "PERÚ VJS": "viajemos_peru",
    "PORTUGAL VJS": "viajemos_portugal",
    "PUERTO RICO VJS": "viajemos_puerto_rico",
    "REPÚBLICA DOMINICANA VJS": "viajemos_republica_dominicana",
    "REINO UNIDO VJS": "viajemos_uk",
    "SALVADOR VJS": "viajemos_salvador",
    "URUGUAY VJS": "viajemos_uruguay",
    "VENEZUELA VJS": "viajemos_venezuela",
}

FLEET_CAR_TYPES = [
    ("economicos", "Economicos",
     "alquiler-autos-economicos", "economy-car-rental", "aluguel-carros-economicos"),
    ("camionetas", "Camionetas (SUV)",
     "alquiler-camionetas", "suv-rental", "aluguel-suv"),
    ("vans", "Vans",
     "alquiler-minivan", "minivan-rental", "aluguel-minivan"),
    ("convertibles", "Convertibles",
     "alquiler-convertibles", "convertible-rental", "aluguel-conversiveis"),
    ("lujo", "Autos de Lujo",
     "alquiler-autos-lujo", "luxury-car-rental", "aluguel-carros-luxo"),
]

TIPO_LP_MAP = {
    "AGENCIAS": "agencia",
    "LOCALIDADES": "localidad",
    "TIPOS DE AUTOS": "tipo_auto",
}

_SENTINELAS = {"no es localidad", "no es agencia", "no es tipo de autos",
               "no es ciudad", "no es oferta"}

DISCLAIMER_PRECIOS = (
    "*Precios basados en los resultados entre los ultimos 12 - 24 meses."
    " Los precios pueden variar segun los dias de alquiler, el tipo de automovil "
    "y la agencia de renta de autos."
)
DISCLAIMER_FINAL = (
    "*Estos precios son sujetos a cambios y variaran dependiendo las fechas "
    "de alquiler, agencia y tipo de vehiculo. Los precios son sugeridos por las "
    "empresas de alquiler y Viajemos no asegura dichos precios. "
    "Verificar precio al momento de la reserva."
)

# ══════════════════════════════════════════════════════════════════════
# SECCIONES MAP POR TIPO
# ══════════════════════════════════════════════════════════════════════

def _build_seccmap_ciudad():
    """VJM CIUDAD: 72 filas (expandido a 17 favoriteCities).
    qs F3-4, sectionCars F5-8, agencies F9-12, rentalCarFaqs F13-23,
    carRental F24-36, favoriteCities F37-72 (17 pares).
    """
    m = {
        3: ("quicksearch", "tit"), 4: ("quicksearch", "desc"),
        5: ("sectionCars", "tit"), 6: ("sectionCars", "desc"),
        7: ("sectionCars", "ip_usa"), 8: ("sectionCars", "ip_bra"),
        9: ("agencies", "tit"), 10: ("agencies", "desc"),
        11: ("agencies", "desc_1"), 12: ("_disclaimer_precios", None),
        13: ("rentalCarFaqs", "tit"), 14: ("rentalCarFaqs", "desc"),
    }
    for i in range(4):
        m[15 + i*2] = ("rentalCarFaqs", f"q_{i+1}")
        m[16 + i*2] = ("rentalCarFaqs", f"faq_{i+1}")
    m[23] = ("_disclaimer_precios", None)
    m[24] = ("carRental", "tit")
    m[25] = ("carRental", "desc")
    for i in range(5):
        m[26 + i*2] = ("carRental", f"tit_{i+1}")
        m[27 + i*2] = ("carRental", f"desc_{i+1}")
    m[36] = ("_disclaimer_final", None)
    m[37] = ("favoriteCities", "tit")
    m[38] = ("favoriteCities", "desc")
    for i in range(17):
        m[39 + i*2] = ("favoriteCities", f"tit_{i+1}")
        m[40 + i*2] = ("favoriteCities", f"desc_{i+1}")
    return m


def _build_seccmap_agencia():
    """VJM AGENCIA: 72 filas (expandido a 17 favoriteCities).
    qs F3-4, sectionCars F5-8, carRental F9-21, rentalCarFaqs F22-32,
    favoriteCities F33-68 (17 pares), agencies F69-72.
    """
    m = {
        3: ("quicksearch", "tit"), 4: ("quicksearch", "desc"),
        5: ("sectionCars", "tit"), 6: ("sectionCars", "desc"),
        7: ("sectionCars", "ip_usa"), 8: ("sectionCars", "ip_bra"),
    }
    m[9] = ("carRental", "tit")
    m[10] = ("carRental", "desc")
    for i in range(5):
        m[11 + i*2] = ("carRental", f"tit_{i+1}")
        m[12 + i*2] = ("carRental", f"desc_{i+1}")
    m[21] = ("_disclaimer_final", None)
    m[22] = ("rentalCarFaqs", "tit")
    m[23] = ("rentalCarFaqs", "desc")
    for i in range(4):
        m[24 + i*2] = ("rentalCarFaqs", f"q_{i+1}")
        m[25 + i*2] = ("rentalCarFaqs", f"faq_{i+1}")
    m[32] = ("_disclaimer_precios", None)
    m[33] = ("favoriteCities", "tit")
    m[34] = ("favoriteCities", "desc")
    for i in range(17):
        m[35 + i*2] = ("favoriteCities", f"tit_{i+1}")
        m[36 + i*2] = ("favoriteCities", f"desc_{i+1}")
    m[69] = ("agencies", "tit")
    m[70] = ("agencies", "desc")
    m[71] = ("agencies", "desc_1")
    m[72] = ("_disclaimer_precios", None)
    return m


def _build_seccmap_localidad():
    """VJM LOCALIDAD: 59 filas (expandido a 17 favoriteCities). SIN carRental.
    qs F3-4, sectionCars F5-8, agencies F9-12,
    favoriteCities F13-48 (17 pares), rentalCarFaqs F49-59.
    """
    m = {
        3: ("quicksearch", "tit"), 4: ("quicksearch", "desc"),
        5: ("sectionCars", "tit"), 6: ("sectionCars", "desc"),
        7: ("sectionCars", "ip_usa"), 8: ("sectionCars", "ip_bra"),
        9: ("agencies", "tit"), 10: ("agencies", "desc"),
        11: ("agencies", "desc_1"), 12: ("_disclaimer_precios", None),
    }
    m[13] = ("favoriteCities", "tit")
    m[14] = ("favoriteCities", "desc")
    for i in range(17):
        m[15 + i*2] = ("favoriteCities", f"tit_{i+1}")
        m[16 + i*2] = ("favoriteCities", f"desc_{i+1}")
    m[49] = ("rentalCarFaqs", "tit")
    m[50] = ("rentalCarFaqs", "desc")
    for i in range(4):
        m[51 + i*2] = ("rentalCarFaqs", f"q_{i+1}")
        m[52 + i*2] = ("rentalCarFaqs", f"faq_{i+1}")
    m[59] = ("_disclaimer_final", None)
    return m


def _build_seccmap_tipo_auto():
    """VJM TIPO_AUTO: 72 filas (expandido a 17 favoriteCities).
    qs F3-4, sectionCars F5-8, agencies F9-12, rentalCarFaqs F13-23,
    favoriteCities F24-59 (17 pares), carRental F60-72.
    """
    m = {
        3: ("quicksearch", "tit"), 4: ("quicksearch", "desc"),
        5: ("sectionCars", "tit"), 6: ("sectionCars", "desc"),
        7: ("sectionCars", "ip_usa"), 8: ("sectionCars", "ip_bra"),
        9: ("agencies", "tit"), 10: ("agencies", "desc"),
        11: ("agencies", "desc_1"), 12: ("_disclaimer_precios", None),
    }
    m[13] = ("rentalCarFaqs", "tit")
    m[14] = ("rentalCarFaqs", "desc")
    for i in range(4):
        m[15 + i*2] = ("rentalCarFaqs", f"q_{i+1}")
        m[16 + i*2] = ("rentalCarFaqs", f"faq_{i+1}")
    m[23] = ("_disclaimer_precios", None)
    m[24] = ("favoriteCities", "tit")
    m[25] = ("favoriteCities", "desc")
    for i in range(17):
        m[26 + i*2] = ("favoriteCities", f"tit_{i+1}")
        m[27 + i*2] = ("favoriteCities", f"desc_{i+1}")
    m[60] = ("carRental", "tit")
    m[61] = ("carRental", "desc")
    for i in range(5):
        m[62 + i*2] = ("carRental", f"tit_{i+1}")
        m[63 + i*2] = ("carRental", f"desc_{i+1}")
    m[72] = ("_disclaimer_final", None)
    return m


SECCIONES_MAP_VJM = {
    "ciudad": _build_seccmap_ciudad(),
    "agencia": _build_seccmap_agencia(),
    "localidad": _build_seccmap_localidad(),
    "tipo_auto": _build_seccmap_tipo_auto(),
}

# Configuracion ImagenesComponentes por tipo (actualizado a 17 favoriteCities)
IMG_COMP_CONFIG = {
    "ciudad":    [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)],
    "agencia":   [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)],
    "localidad": [("agencies", 2), ("favoriteCities", 17)],
    "tipo_auto": [("agencies", 2), ("favoriteCities", 17), ("carRental", 5)],
}


# ══════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def _es_real(val):
    if not val:
        return False
    return val.strip().lower() not in _SENTINELAS


def _slugify(text):
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    return re.sub(r"[-\s]+", "-", text).strip("-")


def _dominio_slug(dominio_bd):
    if not dominio_bd:
        return "viajemos"
    return DOMINIO_BD_TO_SLUG.get(dominio_bd.strip(), "viajemos")


def _pais_slug(dominio_slug):
    if dominio_slug == "viajemos":
        return "estados-unidos"
    return dominio_slug.replace("viajemos_", "").replace("_", "-")


# ══════════════════════════════════════════════════════════════════════
# CONTEXTO LP
# ══════════════════════════════════════════════════════════════════════

def build_ctx(db, lp_id):
    row = db.execute("SELECT * FROM landing_pages WHERE id = ?", (lp_id,)).fetchone()
    if not row:
        raise ValueError(f"LP {lp_id} no encontrada")

    lp = dict(row)
    trinomio = obtener_urls_trinomio(db, lp_id)

    ciudad = lp["ciudad"] if _es_real(lp["ciudad"]) else ""
    localidad = lp["localidad"] if _es_real(lp["localidad"]) else ""
    agencia = lp["agencia"] if _es_real(lp["agencia"]) else ""
    car_category = lp["car_category"] if _es_real(lp["car_category"]) else ""
    tipo_lp_bd = lp["tipo_lp"]

    # Usar url_path (ya es solo el path). Si vacio, extraer de url_completa.
    def _get_path(lang):
        path = trinomio.get(lang, {}).get("url_path", "")
        if path:
            return path
        completa = trinomio.get(lang, {}).get("url_completa", "")
        if completa:
            m = re.search(r"https?://[^/]+(/.*)", completa)
            return m.group(1).lstrip("/") if m else ""
        return ""

    url_es = _get_path("ES")
    url_en = _get_path("EN")
    url_pt = _get_path("PT")

    dominio_slug = _dominio_slug(lp.get("dominio", ""))
    pais_slug = _pais_slug(dominio_slug)

    # tipo_lp_bd='LOCALIDADES' sin localidad real es una CIUDAD pura
    if tipo_lp_bd == "LOCALIDADES" and not localidad:
        tipo_gen = "ciudad"
    else:
        tipo_gen = TIPO_LP_MAP.get(tipo_lp_bd, "ciudad")

    if tipo_lp_bd == "AGENCIAS" and agencia:
        img_prefix = _slugify(f"{agencia} {ciudad}") if ciudad else f"{_slugify(agencia)}-mundo"
    elif tipo_lp_bd == "TIPOS DE AUTOS" and car_category:
        img_prefix = _slugify(f"{car_category} {ciudad}") if ciudad else f"{_slugify(car_category)}-mundo"
    elif tipo_lp_bd == "LOCALIDADES" and localidad:
        # Truncar a max 30 chars para evitar filenames largos
        slug = _slugify(localidad)[:30].rstrip("-")
        img_prefix = slug
    elif ciudad:
        img_prefix = _slugify(ciudad)
    else:
        img_prefix = "default"

    if tipo_lp_bd == "AGENCIAS" and agencia:
        nombre_lp = f"{agencia} {ciudad}" if ciudad else agencia
    elif tipo_lp_bd == "TIPOS DE AUTOS" and car_category:
        nombre_lp = f"{car_category} {ciudad}" if ciudad else car_category
    elif tipo_lp_bd == "LOCALIDADES" and localidad:
        nombre_lp = localidad
    else:
        nombre_lp = ciudad

    return {
        "lp_id": lp_id,
        "lp": lp,
        "tipo_lp_bd": tipo_lp_bd,
        "tipo_gen": tipo_gen,
        "ciudad": ciudad,
        "localidad": localidad,
        "agencia": agencia,
        "car_category": car_category,
        "dominio_bd": lp.get("dominio", ""),
        "dominio_slug": dominio_slug,
        "pais_slug": pais_slug,
        "img_prefix": img_prefix,
        "nombre_lp": nombre_lp,
        "urls": {"ES": url_es, "EN": url_en, "PT": url_pt},
    }


# ══════════════════════════════════════════════════════════════════════
# FILL FUNCTIONS
# ══════════════════════════════════════════════════════════════════════

def fill_landing_page(ws, ctx, db):
    """LandingPage VJM (11 cols A-K)."""
    nombre_lp = ctx["nombre_lp"]
    urls = ctx["urls"]
    dominio = ctx["dominio_bd"]

    ws.cell(row=3, column=1, value="Viajemos")
    ws.cell(row=3, column=2, value=ctx["dominio_slug"])
    ws.cell(row=3, column=3, value=urls.get("ES", ""))
    ws.cell(row=3, column=4, value=VJM_HIGH_PRICE)
    ws.cell(row=3, column=5, value=VJM_LOW_PRICE)

    trinomio = obtener_urls_trinomio(db, ctx["lp_id"])
    col_map = {"EN": (6, 7), "ES": (8, 9), "PT": (10, 11)}
    for lang, (col_t, col_d) in col_map.items():
        url_data = trinomio.get(lang, {}) or {}
        bd_mt = (url_data.get("meta_title") or "").strip()
        bd_md = (url_data.get("meta_description") or "").strip()

        meta_t = bd_mt if bd_mt else (
            generar_meta_titulo(nombre_lp, ctx["tipo_lp_bd"], lang, dominio,
                                precio=VJM_LOW_PRICE, marca="VJM") if nombre_lp else ""
        )
        meta_d = bd_md if bd_md else (
            generar_meta_descripcion(nombre_lp, ctx["tipo_lp_bd"], lang, dominio,
                                     precio=VJM_LOW_PRICE, marca="VJM") if nombre_lp else ""
        )
        ws.cell(row=3, column=col_t, value=meta_t)
        ws.cell(row=3, column=col_d, value=meta_d)


def fill_secciones(ws, ctx, content, tipo_gen,
                   content_en=None, content_pt=None):
    """Hoja Secciones VJM - escribe ES (col C), EN (col D), PT (col E)."""
    mapa = SECCIONES_MAP_VJM.get(tipo_gen, SECCIONES_MAP_VJM["ciudad"])
    content_en = content_en or {}
    content_pt = content_pt or {}

    for row, (bloque, key) in mapa.items():
        if bloque == "_disclaimer_precios":
            ws.cell(row=row, column=3, value=DISCLAIMER_PRECIOS)
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
        elif bloque == "_disclaimer_final":
            ws.cell(row=row, column=3, value=DISCLAIMER_FINAL)
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
        else:
            bloque_data = (content or {}).get(bloque, {})
            val_es = bloque_data.get(key, "") if (bloque_data and key) else ""
            bloque_en = content_en.get(bloque, {})
            bloque_pt = content_pt.get(bloque, {})
            val_en = bloque_en.get(key, "") if (bloque_en and key) else ""
            val_pt = bloque_pt.get(key, "") if (bloque_pt and key) else ""

            # Aplicar rich text (colores inline) a ES, EN, PT
            desc_base_es = bloque_data.get("desc", "") if bloque == "sectionCars" else ""
            desc_base_en = bloque_en.get("desc", "") if bloque == "sectionCars" else ""
            desc_base_pt = bloque_pt.get("desc", "") if bloque == "sectionCars" else ""

            for col, val, lang, base in [
                (3, val_es, "es", desc_base_es),
                (4, val_en, "en", desc_base_en),
                (5, val_pt, "pt", desc_base_pt),
            ]:
                if val and bloque == "sectionCars" and key in ("ip_usa", "ip_bra"):
                    ws.cell(row=row, column=col).value = colorize_fleet_ip(
                        base, val, "vjm", lang)
                elif val:
                    ws.cell(row=row, column=col).value = colorize_text(val, "vjm", lang)
                else:
                    ws.cell(row=row, column=col, value="")


def fill_precios_agencias(ws, ctx):
    """PreciosAgencias VJM (10 agencias, orden TITULOS->HREFs)."""
    ciudad = ctx["ciudad"]  # ciudad REAL, no nombre_lp
    pais = ctx["pais_slug"]
    tipo = ctx["tipo_lp_bd"]
    ciudad_slug = _slugify(ciudad) if ciudad else ""

    for i, (ag_code, ag_display, precio) in enumerate(VJM_AGENCIAS):
        row = 3 + i
        slug = AGENCY_SLUG.get(ag_display, _slugify(ag_display))

        if tipo == "AGENCIAS":
            href_es = f"/es/autos/alquiler-autos-{slug}-{pais}"
            href_en = f"/en/cars/{slug}-car-rental-{pais}"
            href_pt = f"/pt/carros/aluguel-carros-{slug}-{pais}"
        elif tipo in ("LOCALIDADES", "TIPOS DE AUTOS") and ciudad_slug:
            href_es = f"/es/autos/{ciudad_slug}/alquiler-autos-{slug}"
            href_en = f"/en/cars/{ciudad_slug}/{slug}-car-rental"
            href_pt = f"/pt/carros/{ciudad_slug}/aluguel-carros-{slug}"
        else:
            href_es = f"/es/autos/alquiler-autos-{slug}-{pais}"
            href_en = f"/en/cars/{slug}-car-rental-{pais}"
            href_pt = f"/pt/carros/aluguel-carros-{slug}-{pais}"

        # Titulo: usa ciudad REAL. Sin ciudad → solo "Rent a Car"
        tit_en = f"{ag_display} Rent a Car in {ciudad}" if ciudad else f"{ag_display} Rent a Car"
        tit_es = f"{ag_display} Rent a Car en {ciudad}" if ciudad else f"{ag_display} Rent a Car"
        tit_pt = f"{ag_display} Rent a Car em {ciudad}" if ciudad else f"{ag_display} Rent a Car"

        ws.cell(row=row, column=1, value=ag_code)
        ws.cell(row=row, column=2, value=precio)
        ws.cell(row=row, column=3, value=tit_en)
        ws.cell(row=row, column=4, value=tit_es)
        ws.cell(row=row, column=5, value=tit_pt)
        ws.cell(row=row, column=6, value=href_en)
        ws.cell(row=row, column=7, value=href_es)
        ws.cell(row=row, column=8, value=href_pt)


def fill_imagenes_secciones(ws, ctx):
    """ImagenesSecciones VJM (1 fila quicksearch)."""
    prefix = ctx["img_prefix"]
    nombre = ctx["nombre_lp"]
    webp = f"{prefix}-vjm.webp"

    ws.cell(row=3, column=1, value="quicksearch")
    ws.cell(row=3, column=2, value=webp)
    ws.cell(row=3, column=3, value=f"Rent a cheap car in {nombre}")
    ws.cell(row=3, column=4, value=f"Economy car rentals in {nombre}")
    ws.cell(row=3, column=5, value=f"Rentar un auto barato en {nombre}")
    ws.cell(row=3, column=6, value=f"Alquiler de autos economicos en {nombre}")
    ws.cell(row=3, column=7, value=f"Alugue um carro barato em {nombre}")
    ws.cell(row=3, column=8, value=f"Locacao de veiculos economicos em {nombre}")


def fill_imagenes_componentes(ws, ctx, tipo_gen):
    """ImagenesComponentes VJM - segun tipo de LP.

    Escribe filas para (agencies, favoriteCities, carRental) con ORDEN consecutivo.
    """
    prefix = ctx["img_prefix"]
    nombre = ctx["nombre_lp"]
    pais = ctx["pais_slug"]

    sections = IMG_COMP_CONFIG.get(tipo_gen, IMG_COMP_CONFIG["ciudad"])

    row = 3
    for seccion, count in sections:
        for orden in range(1, count + 1):
            if seccion == "agencies":
                webp = f"{prefix}-vjm-agencia-{orden}.webp"
                alt_en = f"Car rental from top agencies in {nombre}"
                tit_en = f"Trusted car rental agencies in {nombre}"
                alt_es = f"Alquiler de autos con mejores agencias en {nombre}"
                tit_es = f"Agencias reconocidas de renta de vehiculos en {nombre}"
                alt_pt = f"Aluguel de carros com melhores locadoras em {nombre}"
                tit_pt = f"Locadoras confiaveis em {nombre}"
                href_es = href_en = href_pt = ""
            elif seccion == "favoriteCities":
                webp = f"{prefix}-vjm-cluster-{orden}.webp"
                alt_en = f"Car rental in nearby city of {nombre}"
                tit_en = f"Car rentals in cities near {nombre}"
                alt_es = f"Alquiler de autos en ciudad cercana a {nombre}"
                tit_es = f"Renta de vehiculos en ciudades cerca de {nombre}"
                alt_pt = f"Aluguel de carros em cidade proxima de {nombre}"
                tit_pt = f"Locacao de veiculos em cidades proximas a {nombre}"
                href_es = href_en = href_pt = ""
            elif seccion == "carRental":
                if orden <= len(FLEET_CAR_TYPES):
                    slug_img, label, h_es_s, h_en_s, h_pt_s = FLEET_CAR_TYPES[orden - 1]
                else:
                    slug_img, label = f"tipo{orden}", f"Tipo {orden}"
                    h_es_s = h_en_s = h_pt_s = ""
                webp = f"{slug_img}-vjm.webp"
                alt_en = f"Rent {label.lower()} in {nombre}"
                tit_en = f"{label} rentals in {nombre}"
                alt_es = f"Alquilar {label.lower()} en {nombre}"
                tit_es = f"Renta de {label.lower()} en {nombre}"
                alt_pt = f"Alugar {label.lower()} em {nombre}"
                tit_pt = f"Aluguel de {label.lower()} em {nombre}"
                href_es = f"/es/autos/{h_es_s}-{pais}" if h_es_s else ""
                href_en = f"/en/cars/{h_en_s}-{pais}" if h_en_s else ""
                href_pt = f"/pt/carros/{h_pt_s}-{pais}" if h_pt_s else ""
            else:
                continue

            ws.cell(row=row, column=1, value=seccion)
            ws.cell(row=row, column=2, value=orden)
            ws.cell(row=row, column=3, value=webp)
            ws.cell(row=row, column=4, value=alt_en)
            ws.cell(row=row, column=5, value=tit_en)
            ws.cell(row=row, column=6, value=alt_es)
            ws.cell(row=row, column=7, value=tit_es)
            ws.cell(row=row, column=8, value=alt_pt)
            ws.cell(row=row, column=9, value=tit_pt)
            ws.cell(row=row, column=10, value=href_en)
            ws.cell(row=row, column=11, value=href_es)
            ws.cell(row=row, column=12, value=href_pt)
            row += 1

    # Limpiar filas residuales hasta max_row
    for r in range(row, ws.max_row + 1):
        for col in range(1, 13):
            ws.cell(row=r, column=col, value="")


# ══════════════════════════════════════════════════════════════════════
# ORQUESTADOR
# ══════════════════════════════════════════════════════════════════════

def _template_for_tipo(tipo_gen):
    return TEMPLATES_BY_TYPE.get(tipo_gen, TEMPLATES_BY_TYPE["ciudad"])


def build_output_filename(ctx):
    fecha = datetime.now().strftime("%d-%m-%Y")
    nombre = ctx["nombre_lp"] or "SinNombre"
    lp_id = ctx["lp_id"]
    return f"{fecha} Cargue de contenido Viajemos {nombre} ID{lp_id}.xlsx"


def build_single(db, lp_id, content=None, content_en=None, content_pt=None,
                 template_path=None, output_dir=OUTPUT_DIR):
    """Construye 1 XLSX VJM completo con traducciones EN/PT."""
    ctx = build_ctx(db, lp_id)
    tipo_gen = ctx["tipo_gen"]
    tmpl = template_path or _template_for_tipo(tipo_gen)

    log.info("Construyendo VJM LP %d: %s (%s) -> dominio=%s plantilla=%s",
             lp_id, ctx["nombre_lp"], ctx["tipo_lp_bd"], ctx["dominio_slug"],
             os.path.basename(tmpl))

    wb = openpyxl.load_workbook(tmpl)
    output_name = build_output_filename(ctx)

    fill_landing_page(wb["LandingPage"], ctx, db)
    fill_precios_agencias(wb["PreciosAgencias"], ctx)
    fill_imagenes_secciones(wb["ImagenesSecciones"], ctx)
    fill_imagenes_componentes(wb["ImagenesComponentes"], ctx, tipo_gen)

    if content:
        fill_secciones(wb["Secciones"], ctx, content, tipo_gen,
                       content_en=content_en, content_pt=content_pt)

    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)
    log.info("Guardado: %s", output_path)
    return output_path


def build_from_json(json_path, lp_ids_map, output_dir=OUTPUT_DIR):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    db = _get_db()
    results = []
    for lp in data:
        if lp.get("marca") != "VJM":
            continue
        tipo = lp["tipo_lp"]
        lp_id = lp_ids_map.get(tipo)
        if not lp_id:
            log.warning("VJM %s: sin lp_id mapeado, saltando", tipo)
            continue
        content = {b["nombre"]: b["contenido"] for b in lp["bloques"]}
        path = build_single(db, lp_id, content=content, output_dir=output_dir)
        results.append(path)
    db.close()
    return results


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--lp-id", type=int)
    ap.add_argument("--json", help="JSON con contenido IA")
    ap.add_argument("--map", help="LP IDs map 'ciudad:111,agencia:222,...'")
    args = ap.parse_args()

    if args.lp_id:
        db = _get_db()
        build_single(db, args.lp_id)
        db.close()
    elif args.json and args.map:
        m = dict(x.split(":") for x in args.map.split(","))
        m = {k: int(v) for k, v in m.items()}
        build_from_json(args.json, m)
    else:
        print("Usar: --lp-id N  o  --json FILE --map ciudad:N,agencia:N,...")
